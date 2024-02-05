# Copyright (c) 2013, Sione Taumoepeau and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import json
import time
import math
import ast
import os.path
import sys
import datetime
import pandas as pd
import xlsxwriter
import openpyxl 
import numpy as np

from openpyxl import load_workbook
from openpyxl.styles import Font
from frappe import _, msgprint, utils
from datetime import datetime, timedelta
from frappe.utils import flt, getdate, datetime, comma_and
from collections import defaultdict
from werkzeug.wrappers import Response
import frappe, erpnext
from collections import defaultdict


def execute(filters=None):
	data = get_data(filters)
	columns = get_columns(filters) if len(data) else []

	return columns, data

def get_columns(filters):
	columns = [
		{
			"label": _("Employee Name"),
			"options": "Employee",
			"fieldname": "employee_name",
			"fieldtype": "Link",
			"width": 160
		},
		{
			"label": _("TIN"),
			"fieldname": "tin",
			"fieldtype": "Data",
			"width": 140
		},
		{
			"label": _("Income Tax Amount"),
			"fieldname": "it_amount",
			"fieldtype": "Currency",
			"options": "currency",
			"width": 140
		},
		{
			"label": _("Gross Pay"),
			"fieldname": "gross_pay",
			"fieldtype": "Currency",
			"options": "currency",
			"width": 140
		},
	]

	return columns

def get_conditions(filters):
	conditions = [""]

#	if filters.get("branch"):
#		conditions.append("branch = '%s' " % (filters["branch"]) )

	if filters.get("company"):
		conditions.append("company = '%s' " % (filters["company"]) )

	if filters.get("month"):
		conditions.append("month(posting_date) = '%s' " % (filters["month"]))

	if filters.get("year"):
		conditions.append("year(posting_date) = '%s' " % (filters["year"]))

	return " and ".join(conditions)


def get_data(filters):

	get_tax = {}
	data = []
	taxdata = []

	fields = ["employee", "tin", "company"]

	employee_details = frappe.get_list("Employee", fields = fields)
	employee_data_dict = {}

	for d in employee_details:
		employee_data_dict.setdefault(
			d.employee,{
				"employee" : d.employee,
				"tin": d.tin,
				"company": d.company
			}
		)

	conditions = get_conditions(filters)

	posting_month = filters.get("month")
	posting_year = filters.get("year")
	company = filters.get("company")
	
	get_gross = frappe.db.sql("""SELECT temp.employee, temp.tin,
		Concat(Ifnull(temp.last_name,' ') ,' ', Ifnull(temp.middle_name,' '),' ', Ifnull(temp.first_name,' ')) as employee_name,
		sal.company, sum(sal.gross_pay) as gross_pay
		FROM `tabEmployee` temp INNER JOIN `tabSalary Slip` sal
		ON temp.employee = sal.employee
		AND sal.docstatus = 1
		AND month(sal.posting_date) = %s
		AND year(sal.posting_date) = %s
		AND sal.company = %s
		GROUP BY temp.employee
		""", (posting_month, posting_year, company), as_dict=1)
	
	employee = {}

	for e in get_gross:
		employee = {
			"emp_id": e.employee,
			"employee_name" : e.employee_name,
			"tin": e.tin,
#			"tin" : employee_data_dict.get(e.employee).get("tin"),
			"it_amount" : 0,
			"gross_pay": e.gross_pay,
			"company": e.company
		}
		data.append(employee)

	get_tax = frappe.db.sql("""SELECT sal.employee, 
		sal.employee_name, 
		sum(ded.amount) as "tax"
		FROM `tabSalary Slip` sal
		INNER JOIN `tabSalary Detail` ded ON
		sal.name = ded.parent
		AND ded.parentfield = 'deductions'
		AND ded.parenttype = 'Salary Slip'
		AND ded.salary_component = "PAYE Tax"
		AND ded.docstatus = 1
		AND ded.amount > 0
		AND month(sal.posting_date) = %s
		AND year(sal.posting_date) = %s
		AND sal.company = %s
		GROUP BY sal.employee
		""", (posting_month, posting_year, company), as_dict=1)


	for e in get_tax:
		employee = {
			"emp_id": e.employee,
			"employee_name" : e.employee_name,
#			"tin" : 0,
			"it_amount" : e.tax,
#			"gross_pay": 0,
			"company": e.company
		}

		data.append(employee)

	combined = defaultdict(dict)
	
	for item in data:
		combined[item['emp_id']].update(item)
	
	taxdata = (list(combined.values()))

	return taxdata

def get_paye_data(posting_month, posting_year, company):
	data = []
	fields = ["employee", "tin", "company"]
	Month = datetime.date(1900, int(posting_month), 1).strftime('%B')

	employee_details = frappe.get_list("Employee", fields = fields)
	employee_data_dict = {}

	for d in employee_details:
		employee_data_dict.setdefault(
			d.employee,{
				"tin": d.tin,
				"employee" : d.employee,
				"date_of_payment" : '',
				"pay_period": 'Fortnightly',
				"company": d.company
			}
		)


	get_gross = frappe.db.sql("""SELECT temp.employee, temp.tin,
		Concat(Ifnull(temp.last_name,' ') ,' ', Ifnull(temp.middle_name,' '),' ', Ifnull(temp.first_name,' ')) as employee_name,
		sal.company, sum(sal.gross_pay) as gross_pay
		FROM `tabEmployee` temp INNER JOIN `tabSalary Slip` sal
		ON temp.employee = sal.employee
		AND month(sal.posting_date) = %s
		AND year(sal.posting_date) = %s
		AND sal.company = %s
		GROUP BY temp.employee
		""", (posting_month, posting_year, company), as_dict=1)
	
	employee = {}

	for e in get_gross:
		employee = {
			"emp_id": e.employee,
			"tin": e.tin,
#			"tin" : employee_data_dict.get(e.employee).get("tin"),
			"employee_name" : e.employee_name,
			"date_of_payment" : '',
			"pay_period": 'Fortnightly',
			"gross_pay": e.gross_pay,
			"benefit" : 0,
			"it_amount" : 0,
		}
		data.append(employee)

	get_tax = frappe.db.sql("""SELECT sal.employee, 
		sal.employee_name, 
		sum(ded.amount) as "tax"
		FROM `tabSalary Slip` sal
		INNER JOIN `tabSalary Detail` ded ON
		sal.name = ded.parent
		AND ded.parentfield = 'deductions'
		AND ded.parenttype = 'Salary Slip'
		AND ded.salary_component = "PAYE Tax"
		AND ded.docstatus = 1
		AND ded.amount > 0
		AND month(sal.posting_date) = %s
		AND year(sal.posting_date) = %s
		AND sal.company = %s
		GROUP BY sal.employee
		""", (posting_month, posting_year, company), as_dict=1)


	for e in get_tax:
		employee = {
			"emp_id": e.employee,
#			"employee_name" : e.employee_name,
			"date_of_payment" : '',
			"pay_period": "Fortnightly",
			"it_amount" : e.tax
		}
	
		data.append(employee)
	combined = defaultdict(dict)
	
	for item in data:
		combined[item['emp_id']].update(item)
	
	taxdata = (list(combined.values()))
	return taxdata

@frappe.whitelist()
def save_data_to_Excel(month, company, year):
    template_filename = "PAYE.xlsm"
    Month = datetime.date(1900, int(month), 1).strftime('%B')
    Abbr = frappe.db.get_value("Company", company, "abbr")
    tin = frappe.db.get_value("Company", company, "tax_id")

    new_filename = f"{Abbr}-PAYE-{Month}-{year}.xlsm"
    save_path = 'fibs.to/private/files/'
    template_file_path = os.path.join(save_path, template_filename)
    new_file_name = os.path.join(save_path, new_filename)

    ferp = frappe.new_doc("File")
    ferp.file_name = new_filename
    ferp.folder = "Home/PAYE_TAX"
    ferp.is_private = 0
    ferp.file_url = f"/private/files/PAYE_TAX/{new_filename}"

    # Load the template workbook
    template_workbook = openpyxl.load_workbook(template_file_path, read_only=False, keep_vba=True)
    new_sheet = template_workbook['PAYE']

    # Update relevant cells in the template with your data
    new_sheet['C5'] = str(tin)
    new_sheet['C7'] = str(Month)
    new_sheet['E7'] = str(year)
    new_sheet['C9'] = str(company)

    font_style = Font(name='Arial', size=10)

    for cell in ['C5', 'C7', 'E7', 'C9']:
        new_sheet[cell].font = font_style

    paye_data = get_paye_data(month, year, company)
    df = pd.DataFrame(paye_data)

    # Adjust DataFrame if needed
    df = df.drop('emp_id', axis=1)  # Remove 'emp_id' column

#    frappe.msgprint(_("Data {0}").format(df))

    # Start writing data from row 13
    start_row = 13
    start_col = 3

    # Writing data
    for index, row in df.iterrows():
        new_sheet.cell(row=index + start_row, column=start_col, value=row["tin"])
        new_sheet.cell(row=index + start_row, column=start_col + 1, value=row["employee_name"])
        new_sheet.cell(row=index + start_row, column=start_col + 2, value=row["date_of_payment"])
        new_sheet.cell(row=index + start_row, column=start_col + 3, value=row["pay_period"])
        new_sheet.cell(row=index + start_row, column=start_col + 4, value=row["gross_pay"])
        new_sheet.cell(row=index + start_row, column=start_col + 5, value=row["benefit"])
        new_sheet.cell(row=index + start_row, column=start_col + 6, value=row["it_amount"])
        # Use the correct column name based on your data
        # new_sheet.cell(row=index + start_row, column=start_row + 7, value=row["net_amount"])

    # Save the modified template as a new file
    template_workbook.save(new_file_name)

    ferp.save()
    frappe.db.sql('''UPDATE `tabFile` SET file_url = %s WHERE file_name = %s''', ("/files/"+new_filename, new_filename), as_dict=True)
    frappe.msgprint(_("File created - {0}").format(new_filename))
@frappe.whitelist()
def get_years():
	year_list = frappe.db.sql_list(
		"""select distinct YEAR(end_date) from `tabSalary Slip` ORDER BY YEAR(end_date) DESC"""
	)
	if not year_list:
		year_list = [getdate().year]

	return "\n".join(str(year) for year in year_list)